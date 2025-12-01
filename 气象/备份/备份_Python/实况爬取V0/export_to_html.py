# ******** ******** ******** 保证各种语言、特殊符号解码正常 ******** ******** ********
import sys
sys.stdout.reconfigure(encoding='utf-8')

# ******** ******** ******** ******** 导入依赖包 ******** ******** ******** ********
import os
import pandas as pd
import math
import json
import openpyxl

# ******** ******** ******** ******** 读取配置文件 ******** ******** ******** ********
with open('config.json', 'r', encoding='utf-8') as file:
    config = json.load(file)

# ******** ******** ******** ******** 配置 ******** ******** ******** ********
station_list_path = config['station_list_path']
scraped_path = config['scraped_path']
to_js_path = config['ready_data_path']

# ******** ******** ******** ******** 读取文件 ******** ******** ******** ********
# 读取station_list
workbook = openpyxl.load_workbook(station_list_path) #对应某个bendi.xlsx文件
# sheet = workbook['仅供测试']
sheet = workbook['站点信息和记录']
# 写文件 另一种写法 英文字母+数字表示行列
"""
sheet['L2'] = ''
sheet['N5'] = ''
"""
# 读文件
rows = list(sheet.values)
station_list = []
if rows:
    headers = rows[0]
    content_rows = rows[1:]
    station_list = [dict(zip(headers, row)) for row in content_rows]
    station_list = [o for o in station_list if isinstance(o['id'], int)]
else:
    print("The sheet of station list is empty.")


# ******** ******** ******** ******** 把爬取结果写入 station list 文件, 会跳过已有数据单元格 ******** ******** ******** ********
#scraped = [{"wmo":"","min":-18.8,"max":-12.4,"avg":""},{"wmo":"","min":-21.7,"max":-12.8,"avg":""},{"wmo":"","min":-11.9,"max":-9,"avg":""},{"wmo":"","min":-19,"max":-6.6,"avg":""},{"wmo":"","min":-23.2,"max":-8.5,"avg":""},{"wmo":"","min":-18.5,"max":-9.9,"avg":""},{"wmo":"","min":-23.1,"max":-13.3,"avg":""},{"wmo":44212,"min":-16.9,"max":-5.2,"avg":""},{"wmo":44221,"min":-17,"max":-8.8,"avg":""},{"wmo":44224,"min":-8,"max":-1.2,"avg":""},{"wmo":44225,"min":-16.4,"max":-6.2,"avg":""},{"wmo":44203,"min":-22.2,"max":-6.2,"avg":""},{"wmo":44292,"min":-20.1,"max":-8.8,"avg":""},{"wmo":"","min":-28.2,"max":-10.9,"avg":""},{"wmo":44291,"min":-28.5,"max":-11,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-27,"max":-15.6,"avg":""},{"wmo":44275,"min":"","max":"","avg":""},{"wmo":44284,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44265,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44277,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-25.1,"max":-12,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44207,"min":"","max":"","avg":""},{"wmo":44231,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-25,"max":-13.1,"avg":""},{"wmo":"","min":-27.2,"max":-13.6,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44215,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44213,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44263,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":44229,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-26.8,"max":-15.1,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":31702,"min":-8.7,"max":-6.7,"avg":""},{"wmo":31532,"min":-17.5,"max":-10.9,"avg":""},{"wmo":31478,"min":-20.5,"max":-10.4,"avg":""},{"wmo":31329,"min":-21.4,"max":-13.5,"avg":""},{"wmo":31348,"min":-29.9,"max":-16.9,"avg":""},{"wmo":30673,"min":-30.8,"max":-18.3,"avg":""},{"wmo":30565,"min":-40.3,"max":-24.6,"avg":""},{"wmo":30664,"min":-37.4,"max":-19.1,"avg":""},{"wmo":30636,"min":-19.1,"max":-13,"avg":""},{"wmo":30622,"min":-18.5,"max":-10.8,"avg":""},{"wmo":36104,"min":-10.6,"max":-6.4,"avg":""},{"wmo":36096,"min":-10.4,"max":-7.5,"avg":""},{"wmo":36307,"min":-16.5,"max":-6.9,"avg":""},{"wmo":36259,"min":-15.8,"max":-6,"avg":""},{"wmo":30781,"min":-32.3,"max":-13.1,"avg":""},{"wmo":36535,"min":-14.3,"max":-2.9,"avg":""},{"wmo":36566,"min":-15.1,"max":-9.1,"avg":""},{"wmo":47005,"min":-10.1,"max":-2.8,"avg":""},{"wmo":54342,"min":-8.5,"max":6.6,"avg":""},{"wmo":54252,"min":-14.7,"max":2.8,"avg":""},{"wmo":54161,"min":-12.6,"max":-1.7,"avg":""},{"wmo":54273,"min":"","max":"","avg":""},{"wmo":50953,"min":-12.6,"max":-3.6,"avg":""},{"wmo":50673,"min":-13.5,"max":-5.6,"avg":""},{"wmo":50353,"min":-25.9,"max":-12.8,"avg":""},{"wmo":50246,"min":-29.1,"max":-13.4,"avg":""},{"wmo":50247,"min":"","max":"","avg":""},{"wmo":50136,"min":-30.9,"max":-14.8,"avg":""},{"wmo":50137,"min":-35.4,"max":-21.8,"avg":""},{"wmo":50431,"min":-30,"max":-13,"avg":""},{"wmo":50434,"min":-31.7,"max":-13.4,"avg":""},{"wmo":50425,"min":-24.1,"max":-14.4,"avg":""},{"wmo":50524,"min":"","max":"","avg":""},{"wmo":50527,"min":-21.6,"max":-13.9,"avg":""},{"wmo":50525,"min":"","max":"","avg":""},{"wmo":50526,"min":"","max":"","avg":""},{"wmo":50727,"min":-25.4,"max":-9.8,"avg":""},{"wmo":53392,"min":-19.3,"max":-1.3,"avg":""},{"wmo":55294,"min":-18.9,"max":-0.7,"avg":""},{"wmo":56034,"min":-28.7,"max":-4.8,"avg":""},{"wmo":52908,"min":-24.8,"max":-4.9,"avg":""},{"wmo":52323,"min":-1.8,"max":8.1,"avg":""},{"wmo":52101,"min":-7.8,"max":8.5,"avg":""},{"wmo":51186,"min":-21.3,"max":-7.4,"avg":""},{"wmo":51076,"min":-8.2,"max":-0.6,"avg":""},{"wmo":51573,"min":-2.3,"max":6.5,"avg":""},{"wmo":51463,"min":-8.9,"max":4,"avg":""},{"wmo":51542,"min":-28.6,"max":-16.5,"avg":""},{"wmo":51437,"min":"","max":"","avg":""},{"wmo":53463,"min":-14.5,"max":-1.4,"avg":""},{"wmo":54511,"min":1.1,"max":7.9,"avg":""},{"wmo":54517,"min":1.8,"max":7.8,"avg":""},{"wmo":53614,"min":"","max":"","avg":""},{"wmo":53698,"min":"","max":"","avg":""},{"wmo":53772,"min":"","max":"","avg":""},{"wmo":52866,"min":"","max":"","avg":""},{"wmo":54823,"min":"","max":"","avg":""},{"wmo":57083,"min":"","max":"","avg":""},{"wmo":57131,"min":"","max":"","avg":""},{"wmo":57245,"min":"","max":"","avg":""},{"wmo":57687,"min":"","max":"","avg":""},{"wmo":58606,"min":7.3,"max":19.6,"avg":""},{"wmo":57494,"min":3.2,"max":18.2,"avg":""},{"wmo":58321,"min":"","max":"","avg":""},{"wmo":58238,"min":"","max":"","avg":""},{"wmo":58457,"min":6.6,"max":16,"avg":""},{"wmo":58646,"min":4.5,"max":19,"avg":""},{"wmo":58847,"min":"","max":"","avg":""},{"wmo":59287,"min":"","max":"","avg":""},{"wmo":59758,"min":17.9,"max":22.7,"avg":""},{"wmo":59948,"min":15.9,"max":21.9,"avg":""},{"wmo":59838,"min":18.7,"max":24,"avg":""},{"wmo":59431,"min":"","max":"","avg":""},{"wmo":56966,"min":"","max":"","avg":""},{"wmo":56778,"min":5.6,"max":15,"avg":""},{"wmo":57816,"min":4.7,"max":16.7,"avg":""},{"wmo":57516,"min":"","max":"","avg":""},{"wmo":57413,"min":"","max":"","avg":""},{"wmo":56187,"min":1.9,"max":17.3,"avg":""},{"wmo":55591,"min":-4.8,"max":9.5,"avg":""},{"wmo":52267,"min":"","max":"","avg":""},{"wmo":53478,"min":-16.5,"max":2.2,"avg":""},{"wmo":58367,"min":8.8,"max":13.5,"avg":""},{"wmo":58826,"min":"","max":"","avg":""},{"wmo":58623,"min":"","max":"","avg":""},{"wmo":57778,"min":"","max":"","avg":""},{"wmo":52889,"min":-5,"max":13.2,"avg":""},{"wmo":50341,"min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-33,"max":-16,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-33.1,"max":-13,"avg":""},{"wmo":"","min":-17.1,"max":-3.1,"avg":""},{"wmo":"","min":-23.1,"max":-8.8,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":-23.9,"max":-5.1,"avg":""},{"wmo":"","min":-16.1,"max":-7.1,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":58437,"min":"","max":"","avg":""},{"wmo":"","min":-20.5,"max":-11.2,"avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":"","min":"","max":"","avg":""},{"wmo":58968,"min":"","max":"","avg":""},{"wmo":45007,"min":"","max":"","avg":""},{"wmo":45011,"min":"","max":"","avg":""},{"wmo":31137,"min":"","max":"","avg":""},{"wmo":24688,"min":-49.6,"max":-43.8,"avg":""},{"wmo":24585,"min":-50.8,"max":-49.3,"avg":""},{"wmo":24588,"min":-50.4,"max":-46.8,"avg":""},{"wmo":24691,"min":-43.4,"max":-40.6,"avg":""},{"wmo":24266,"min":-49,"max":-46.6,"avg":""},{"wmo":24684,"min":-49.6,"max":-44.3,"avg":""},{"wmo":24382,"min":-45.3,"max":-42.6,"avg":""},{"wmo":24959,"min":-28.1,"max":-25.6,"avg":""},{"wmo":24477,"min":"","max":"","avg":""},{"wmo":25428,"min":-25.5,"max":-17.6,"avg":""},{"wmo":25700,"min":-44.8,"max":-33.6,"avg":""},{"wmo":24507,"min":-44.7,"max":-40,"avg":""},{"wmo":"","min":-36.8,"max":-22,"avg":""},{"wmo":"","min":-29.4,"max":-13.4,"avg":""},{"wmo":"","min":-36.8,"max":-20.2,"avg":""}]
with open(scraped_path, 'r', encoding='utf-8') as f:
    scraped = json.load(f)['scraped_data']
# print(scraped)
def write_into_station_list():
    c1 = 15 # station_list: min-15, max=16, avg=17
    # temp_count = 0
    for i, o in enumerate(station_list): # index before element
        #超长判断语句
        is_writable_row = (o['USAF'] is not None) and (o['use'] is not None) and (o['min'] is None) and (o['max'] is None) and (o['avg'] is None) and ( str(scraped[i]['min']) != '' or str(scraped[i]['max']) != '' )
        # temp_count += 1
        # print(temp_count, is_writable_row)
        if is_writable_row:  
            sheet.cell(row = i + 2, column = c1, value = scraped[i]['min'])
            sheet.cell(row = i + 2, column = c1 + 1, value = scraped[i]['max'])
            sheet.cell(row = i + 2, column = c1 + 2, value = scraped[i]['avg'])
    workbook.save(station_list_path) #save the file

# ******** ******** ******** ******** 把 station list 数据写入 .js 文件 供前端使用 ******** ******** ******** ********
# 筛选出有气温数据的站
filteredArr = [e for e in station_list if isinstance(e['id'], int) and not ( (e['min'] is None) and (e['max'] is None) )]
# 特供函数
def sort_mix_type(e, attr, asc):
    strArr = []
    for e1 in headers:
        strArr.append(e1)
    for s in strArr:
        if isinstance(e[s], int):
            e[s] = float(e[s])
        if e[s] is None: # 如果单元格是空值，那么把单元格的值设为无限大或无限小float数字类型
            e[s] = float('inf') if asc else float('-inf')
    return e[attr]
# 数组排序，先按低温，再按高温
sortedArr = sorted(
    filteredArr,
    key = lambda x: (sort_mix_type(x, 'min', True), sort_mix_type(x, 'max', True))
)
for o in sortedArr:
    for k in o:
        # print(o[k], type(o[k]))
        if isinstance(o[k], float) and math.isinf(o[k]): # 清空 -999的值(之后务必参考bad_list检查异常站)
            o[k] = ''
        if k == 'prefix' and isinstance(o[k], str) and len(o[k]) > 0: # 合并prefix和cn_name文本
            o['cn_name'] = o['prefix'] + o['cn_name']
# 定制导出东亚json
def arr_east_asia():
    eaArr = [e for e in sortedArr if e['east_asia'] == 'Y']
    return eaArr

# 把多个定制的对象数组导出到js文件
def export_to_JS():
    json_all_stations = json.dumps(sortedArr, indent=4) #全部有数据的站
    json_ea_stations = json.dumps(arr_east_asia(), indent=4) # 东亚json
    js_all_stations = f'export const array_of_stations_by_different_filters = {{\n"汇总": {json_all_stations}, \n"东亚": {json_ea_stations}}};'
    with open(to_js_path, "w", encoding="utf-8") as f:
        f.write(js_all_stations)

    print(f"Successfully created .js file")

# ******** ******** ******** ******** 运行 ******** ******** ******** ********
write_into_station_list() # 1.把爬取结果写入 station list 文件。如果报错 "IndexError: list index out of range"，检查表格sheet，可能是"仅供测试"的sheet
export_to_JS() # 2.把 station list data 导出到 js 文件 供前端使用