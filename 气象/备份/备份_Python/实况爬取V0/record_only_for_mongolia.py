# ******** ******** ******** 保证各种语言、特殊符号解码正常 ******** ******** ********
import sys
sys.stdout.reconfigure(encoding='utf-8')

# ******** ******** ******** ******** 导入依赖包 ******** ******** ******** ********
import os
import json
import openpyxl
import time
from datetime import datetime, timedelta, timezone

# ******** ******** ******** ******** 读取配置文件 ******** ******** ******** ********
with open('config.json', 'r', encoding='utf-8') as file:
    config = json.load(file)

# ******** ******** ******** ******** 配置 ******** ******** ******** ********
station_list_path = config['station_list_path']
mongolia_json_path = config['only_for_mongolia_json_path']
mongolia_records_path = os.path.abspath(config['mongolia_records']['path'])
ws_names = config['mongolia_records']['sheet_names']

# ******** ******** ******** ******** 读取 station_list 文件 ******** ******** ******** ********
workbook = openpyxl.load_workbook(station_list_path)
sheet = workbook['站点信息和记录']
rows = list(sheet.values)
station_list = []
if rows:
    headers = rows[0]
    content_rows = rows[1:]
    station_list = [dict(zip(headers, row)) for row in content_rows]
    mongolia_list = [o for o in station_list if o['id'] is not None and int(o['id']) >= 1 and int(o['id']) <= 13] # 读取仅蒙古站点数据
else:
    print("The sheet of station list is empty.")
#把数组min, max, avg值为None的替换为单字"无"
for o in mongolia_list:
    for k in o:
        if k != 'min' and k != 'max' and k != 'avg':
            if o[k] is None:
                o[k] = ''
        else:
            if o[k] is None:
                o[k] = '无'

# ******** ******** ******** ******** 导出为json文件 ******** ******** ******** ********
json_string = f'{{\n"mongolia_list":\n{json.dumps(mongolia_list, indent=4)}\n}}' #全部蒙古的站
with open(mongolia_json_path, "w", encoding="utf-8") as f:
    f.write(json_string)
    print(f"Successfully created only for mongolia .json file")

# ******** ******** ******** ******** 打开 蒙古存档.xlsx文件 并写入蒙古13个冷站数据 ******** ******** ******** ********
#读取某个sheet
workbook_mn = openpyxl.load_workbook(mongolia_records_path)
sheet_mongolia_min = workbook_mn[ws_names[0]]
sheet_mongolia_max = workbook_mn[ws_names[1]]

# 获取、准备日期、时间字符串
def get_utc_datetime_parts():
    now = datetime.now(timezone.utc)
    return {
        "YYYY": now.strftime("%Y"),
        "MM": now.strftime("%m"),
        "DD": now.strftime("%d"),
        "hh": now.strftime("%H"),
        # "mm": now.strftime("%M") 分钟
    }
dt = get_utc_datetime_parts()
# target_date = '2025/11/27'
target_date = dt['YYYY'] + '/' + dt['MM'] + '/' + dt['DD'] # 当前UTC时区日期

has_the_date = False
sheet_name = ''
def fill_value(page):
    global sheet_name
    sheet_name = ws_names[0] if page == 'min' else ws_names[1]

    # 读: 低温存档
    sheet = sheet_mongolia_min if page == 'min' else sheet_mongolia_max
    rows = list(sheet.values)
    filtered_rows = []
    if rows:
        headers = rows[0]
        content_rows = rows[1:]
        filtered_rows = [dict(zip(headers, row)) for row in content_rows]
        filtered_rows = [o for o in filtered_rows if o['日期'] is not None] # 对应表格等同于去除末尾可能有的大量空行
    else:
        print(f'The sheet({sheet_name}) is empty.')
        
    # 读取 刚才导出的 only_for_mongolia json 文件
    with open(mongolia_json_path, 'r', encoding='utf-8') as f:
        mongolia_arr = json.load(f)['mongolia_list']

    global has_the_date
    for i, o in enumerate(filtered_rows):
        date_cell = sheet.cell(row = i + 2, column = 1)
        if date_cell.value.strftime('%Y/%m/%d') == target_date: # 如果表格中已存在某个日期字符串
            has_the_date = True
            if o['东戈壁'] is not None: # 如果该日期已填数据
                print(f'{sheet_name}： ALREADY  filled for the date(---- {target_date} ----) before, Check if the date is you want!!!')
            else:
                for index, ele in enumerate(mongolia_arr):
                    sheet.cell(row = i + 2, column = index + 2, value = ele['min'] if page == 'min' else ele['max'])
                workbook_mn.save(mongolia_records_path) #save the file
                print(f'{sheet_name}: SUCCESSFULLY  filled data for the date (---- {target_date} ----)!')

fill_value('min')
fill_value('max')
if not has_the_date:
    print(f'Warning: Fill more date into the sheet({sheet_name}) for the first and then come back here to fill data!!!')